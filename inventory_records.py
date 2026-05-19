from copy import copy
from dataclasses import dataclass
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

DATE_COLUMNS = ["Date In", "Death Date", "Shipped out date"]
RECORD_LOOKBACK_DAYS = 30

RECORD_COLUMNS = [
    ("Date", 20),
    ("Prev. Inventory", 22),
    ("Entries", 18),
    ("Deads", 18),
    ("Shipped", 18),
    ("Inventory", 20),
]


@dataclass(frozen=True)
class InventoryRecordSheet:
    sheet_title: str
    section_title: str
    table_title: str
    records: pd.DataFrame


def normalize_locations(df: pd.DataFrame, no_location_label: str) -> pd.Series:
    return df["Location"].fillna(no_location_label).astype(str).str.strip().replace("", no_location_label)


def filter_owned_inventory(df: pd.DataFrame, ownership: str = "Brandao Cattle") -> pd.DataFrame:
    return df.loc[df["Ownership"] == ownership].copy()


def prepare_record_dates(df: pd.DataFrame) -> pd.DataFrame:
    record_df = df.copy()

    for column_name in DATE_COLUMNS:
        record_df[column_name] = pd.to_datetime(record_df[column_name], errors="coerce").dt.normalize()

    # Shipped rows without an exit date still need to leave the running inventory.
    missing_shipped_date = (record_df["Status"] == "Shipped") & record_df["Shipped out date"].isna()
    record_df.loc[missing_shipped_date, "Shipped out date"] = record_df.loc[missing_shipped_date, "Date In"]

    return record_df


def count_events_by_day(date_series: pd.Series, column_name: str) -> pd.Series:
    clean_dates = date_series.dropna()

    if clean_dates.empty:
        return pd.Series(dtype="int64", name=column_name)

    return clean_dates.groupby(clean_dates).size().rename(column_name)


def build_inventory_record(df: pd.DataFrame) -> pd.DataFrame:
    record_df = prepare_record_dates(df)

    entries = count_events_by_day(record_df["Date In"], "entries")
    deaths = count_events_by_day(record_df.loc[record_df["Status"] == "Dead", "Death Date"], "deaths")
    shipped = count_events_by_day(record_df.loc[record_df["Status"] == "Shipped", "Shipped out date"], "shipped")

    movements = pd.concat([entries, deaths, shipped], axis=1).sort_index()

    if movements.empty:
        return pd.DataFrame(columns=["date", "prev_inventory", "entries", "deaths", "shipped", "inventory"])

    daily_index = pd.date_range(movements.index.min(), movements.index.max(), freq="D")
    movements = movements.reindex(daily_index).fillna(0).astype(int)
    movements.index.name = "date"

    cumulative = movements.cumsum()
    inventory = cumulative["entries"] - cumulative["deaths"] - cumulative["shipped"]

    report = movements.copy()
    report["inventory"] = inventory
    report["prev_inventory"] = report["inventory"].shift(1).fillna(0).astype(int)
    report = report.reset_index()
    latest_date = report["date"].max()
    start_date = latest_date - pd.Timedelta(days=RECORD_LOOKBACK_DAYS - 1)
    report = report.loc[report["date"] >= start_date].reset_index(drop=True)

    return report[["date", "prev_inventory", "entries", "deaths", "shipped", "inventory"]]


def build_sheet_title(base_title: str, used_titles: set[str]) -> str:
    clean_title = re.sub(r"[\[\]:*?/\\]", " ", f"{base_title} Record").strip()
    clean_title = re.sub(r"\s+", " ", clean_title) or "Inventory Record"
    sheet_title = clean_title[:31]

    counter = 2
    while sheet_title in used_titles:
        suffix = f" {counter}"
        sheet_title = f"{clean_title[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_titles.add(sheet_title)
    return sheet_title


def build_inventory_record_sheets(
    ranch_dataframes: dict[str, pd.DataFrame],
    ranch_section_titles: dict[str, str],
    california_ranch_name: str,
    no_location_label: str,
    california_locations: list[str] | None = None,
) -> list[InventoryRecordSheet]:
    record_sheets = []
    used_titles = set()

    for ranch_name, ranch_df in ranch_dataframes.items():
        owned_df = filter_owned_inventory(ranch_df)
        section_title = ranch_section_titles.get(ranch_name, ranch_name)

        if ranch_name == california_ranch_name:
            location_df = owned_df.copy()
            location_df["Location"] = normalize_locations(location_df, no_location_label)
            locations = california_locations or sorted(location_df["Location"].dropna().unique())

            for location in locations:
                records = build_inventory_record(location_df.loc[location_df["Location"] == location])
                record_sheets.append(
                    InventoryRecordSheet(
                        sheet_title=build_sheet_title(location, used_titles),
                        section_title=section_title,
                        table_title=location,
                        records=records,
                    )
                )
        else:
            record_sheets.append(
                InventoryRecordSheet(
                    sheet_title=build_sheet_title(ranch_name, used_titles),
                    section_title=section_title,
                    table_title=ranch_name,
                    records=build_inventory_record(owned_df),
                )
            )

    return record_sheets


def apply_record_sheet_styles(ws) -> None:
    for column_letter, (_, width) in zip("ABCDEF", RECORD_COLUMNS):
        ws.column_dimensions[column_letter].width = width

    ws.sheet_view.showGridLines = False
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )


def write_record_headers(ws) -> None:
    ws["A1"] = "BRANDAO CATTLE"
    ws["A1"].font = Font(bold=True, size=15)

    ws["A2"] = "INVENTORY RECORDS"
    ws["A2"].font = Font(size=13)

    ws["A3"] = '="DATE: " & TEXT(TODAY(), "mm/dd/yyyy")'
    ws["A3"].font = Font(size=12)
    ws["B3"] = None


def write_record_table_header(ws, row_number: int, table_title: str) -> int:
    thin_gray = Side(style="thin", color="808080")

    ws.cell(row=row_number, column=1, value=table_title).font = Font(bold=False, size=13)

    header_row = row_number + 1

    for column_index, (header, _) in enumerate(RECORD_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="left" if header == "Date" else "center")
        cell.border = Border(bottom=thin_gray)

    return header_row


def write_record_rows(ws, start_row: int, records: pd.DataFrame) -> int:
    current_row = start_row

    for values in records.itertuples(index=False):
        ws.cell(row=current_row, column=1, value=values.date)
        ws.cell(row=current_row, column=2, value=int(values.prev_inventory))
        ws.cell(row=current_row, column=3, value=int(values.entries))
        ws.cell(row=current_row, column=4, value=int(values.deaths))
        ws.cell(row=current_row, column=5, value=int(values.shipped))
        ws.cell(row=current_row, column=6, value=int(values.inventory))
        current_row += 1

    return current_row


def format_record_rows(ws, first_row: int, last_row: int) -> None:
    if last_row < first_row:
        return

    row_separator = Side(style="thin", color="E6E6E6")

    for row in range(first_row, last_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=1).number_format = "mm/dd/yyyy"

        for column_index in range(2, 7):
            ws.cell(row=row, column=column_index).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=column_index).number_format = "#,##0"

        ws.cell(row=row, column=3).number_format = '+ #,##0;+ #,##0;"0"'
        ws.cell(row=row, column=4).number_format = '- #,##0;- #,##0;"0"'
        ws.cell(row=row, column=5).number_format = '- #,##0;- #,##0;"0"'

        for column_index in range(1, len(RECORD_COLUMNS) + 1):
            ws.cell(row=row, column=column_index).border = Border(bottom=row_separator)


def bold_last_inventory_value(ws, data_start_row: int, data_end_row: int) -> None:
    if data_end_row < data_start_row:
        return

    ws.cell(row=data_end_row, column=6).font = Font(bold=True)


def apply_vertical_centering(ws, last_row: int, last_column: int) -> None:
    for row_number in range(1, last_row + 1):
        for column_index in range(1, last_column + 1):
            cell = ws.cell(row=row_number, column=column_index)
            alignment = copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment


def apply_record_print_layout(ws, last_row: int) -> None:
    for row_number in range(1, last_row + 1):
        ws.row_dimensions[row_number].height = 18

    apply_vertical_centering(ws, last_row, len(RECORD_COLUMNS))
    ws.print_area = f"A1:F{last_row}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def write_inventory_record_sheet(workbook, record_sheet: InventoryRecordSheet) -> None:
    worksheet = workbook.create_sheet(record_sheet.sheet_title)

    apply_record_sheet_styles(worksheet)
    write_record_headers(worksheet)

    worksheet.cell(row=6, column=1, value=record_sheet.section_title).font = Font(bold=True, size=13)

    header_row = write_record_table_header(worksheet, 8, record_sheet.table_title)
    data_start_row = header_row + 1
    next_row = write_record_rows(worksheet, data_start_row, record_sheet.records)
    data_end_row = next_row - 1

    format_record_rows(worksheet, data_start_row, data_end_row)
    bold_last_inventory_value(worksheet, data_start_row, data_end_row)
    apply_record_print_layout(worksheet, max(data_end_row, header_row))
