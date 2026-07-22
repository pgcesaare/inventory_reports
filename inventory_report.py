from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border, Font, Side

from inventory_records import InventoryRecordSheet, build_inventory_record_sheets, write_inventory_record_sheet

# Ruta principal en Windows y ruta secundaria actual.
BASE_PATH_CANDIDATES = [
    Path("C:/Users/cesar/OneDrive/Documentos"),
    Path("/Users/pgcesaare/OneDrive/Documentos"),
]

RANCH_FILES = {
    "California Inventory": "California Inventory.xlsx",
    "La Esperanza Ranch": "Inventory at Dominguez - Guess Cattle.xlsx",
    "Cesar Frias Ranch": "Inventory at Frias - Guess Cattle.xlsx",
    "Fullmer Cattle": "Inventory at Fullmer Cattle.xlsx",
}

CALIFORNIA_RANCH_NAME = "California Inventory"
NO_LOCATION_LABEL = "No Location"
RANCH_SECTION_TITLES = {
    CALIFORNIA_RANCH_NAME: "California",
    "La Esperanza Ranch": "Washington",
    "Cesar Frias Ranch": "Idaho",
    "Fullmer Cattle": "Kansas",
}

COLUMNS = [
    ("Breed", 32),
    ("Quantity", 15),
    ("Avg. Price", 15),
    ("Avg. DOF", 15),
    ("Min Date", 15),
    ("Max Date", 15),
    ("Total", 18),
]

InventoryAssignment = pd.DataFrame | dict[str, pd.DataFrame]


def resolve_base_path() -> Path:
    required_files = list(RANCH_FILES.values())

    for base_path in BASE_PATH_CANDIDATES:
        if base_path.exists() and all((base_path / filename).exists() for filename in required_files):
            return base_path

    missing_by_path = []

    for base_path in BASE_PATH_CANDIDATES:
        missing_files = [filename for filename in required_files if not (base_path / filename).exists()]
        missing_by_path.append(f"{base_path}: {', '.join(missing_files)}")

    missing_detail = " | ".join(missing_by_path)
    raise FileNotFoundError(f"No se encontraron todos los archivos de inventario. Revisado en: {missing_detail}")


BASE_PATH = resolve_base_path()
OUTPUT_DIR = BASE_PATH / "Inventory Reports"


def load_ranch_file(filename: str) -> pd.DataFrame:
    file_path = BASE_PATH / filename
    try:
        return pd.read_excel(file_path)
    except PermissionError as exc:
        raise PermissionError(
            f"No se pudo leer '{file_path}'. "
            "Windows nego el acceso al archivo. "
            "Cierra el archivo en Excel, espera a que OneDrive termine de sincronizar "
            "y asegurate de que el archivo este disponible localmente."
        ) from exc


def load_ranch_dataframes() -> dict[str, pd.DataFrame]:
    return {ranch_name: load_ranch_file(filename) for ranch_name, filename in RANCH_FILES.items()}


def filter_inventory(df: pd.DataFrame) -> pd.DataFrame:
    mask = (df["Ownership"] == "Brandao Cattle") & (df["Status"] == "Feeding")
    return df.loc[mask].copy()


def build_inventory(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby(by="Breed")
        .agg(
            quantity=("Breed", "size"),
            avg_price=("Purchase Price", "mean"),
            avg_DOF=("DOF", "mean"),
            min_date=("Date In", "min"),
            max_date=("Date In", "max"),
            total=("Purchase Price", "sum"),
        )
        .sort_index()
    )

    summary["avg_DOF"] = summary["avg_DOF"].round().astype(int)

    return summary


def build_inventory_by_location(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    if df.empty:
        return {NO_LOCATION_LABEL: build_inventory(df)}

    location_df = df.copy()
    location_df["Location"] = location_df["Location"].fillna(NO_LOCATION_LABEL)
    location_df["Location"] = location_df["Location"].astype(str).str.strip().replace("", NO_LOCATION_LABEL)

    inventories = {}

    for location, location_group_df in location_df.groupby(by="Location", sort=True):
        inventories[location] = build_inventory(location_group_df)

    return inventories


def build_inventory_assignments(ranch_dataframes: dict[str, pd.DataFrame]) -> dict[str, InventoryAssignment]:
    inventories = {}

    for ranch_name, ranch_df in ranch_dataframes.items():
        filtered_df = filter_inventory(ranch_df)
        if ranch_name == CALIFORNIA_RANCH_NAME:
            inventories[ranch_name] = build_inventory_by_location(filtered_df)
        else:
            inventories[ranch_name] = build_inventory(filtered_df)

    return inventories


def load_inventory_assignments() -> dict[str, InventoryAssignment]:
    return build_inventory_assignments(load_ranch_dataframes())


def build_output_path() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    report_date = pd.Timestamp.today().strftime("%m.%d.%Y")
    filename = f"BRANDAO CATTLE INVENTORY REPORT {report_date}.xlsx"
    return OUTPUT_DIR / filename


def apply_sheet_styles(ws) -> None:
    for column_letter, (_, width) in zip("ABCDEFG", COLUMNS):
        ws.column_dimensions[column_letter].width = width

    ws.sheet_view.showGridLines = False
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )


def write_headers(ws) -> None:
    ws["A1"] = "BRANDAO CATTLE"
    ws["A1"].font = Font(bold=True, size=15)

    ws["A2"] = "INVENTORY REPORT"
    ws["A2"].font = Font(size=13)

    ws["A3"] = '="DATE: " & TEXT(TODAY(), "mm/dd/yyyy")'
    ws["A3"].font = Font(size=12)
    ws["B3"] = None


def write_table_header(ws, row_number: int, ranch_name: str) -> int:
    thin_gray = Side(style="thin", color="808080")

    ws.cell(row=row_number, column=1, value=ranch_name).font = Font(bold=False, size=13)

    header_row = row_number + 1

    for column_index, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True, size=12)
        if header == "Breed":
            alignment = "left"
        elif header == "Total":
            alignment = "right"
        else:
            alignment = "center"
        cell.alignment = Alignment(horizontal=alignment)
        cell.border = Border(bottom=thin_gray)

    return header_row


def write_inventory_rows(ws, start_row: int, inventory_df: pd.DataFrame) -> int:
    current_row = start_row

    for breed, values in inventory_df.iterrows():
        ws.cell(row=current_row, column=1, value=breed)
        ws.cell(row=current_row, column=2, value=int(values["quantity"]))
        ws.cell(row=current_row, column=3, value=float(values["avg_price"]))
        ws.cell(row=current_row, column=4, value=int(values["avg_DOF"]))
        ws.cell(row=current_row, column=5, value=values["min_date"])
        ws.cell(row=current_row, column=6, value=values["max_date"])
        ws.cell(row=current_row, column=7, value=float(values["total"]))
        current_row += 1

    return current_row


def format_data_rows(ws, first_row: int, last_row: int) -> None:
    if last_row < first_row:
        return

    row_separator = Side(style="thin", color="E6E6E6")

    for row in range(first_row, last_row + 1):
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right")

        ws.cell(row=row, column=2).number_format = "#,##0"
        ws.cell(row=row, column=3).number_format = "$#,##0.00"
        ws.cell(row=row, column=4).number_format = "0"
        ws.cell(row=row, column=5).number_format = "mm/dd/yyyy"
        ws.cell(row=row, column=6).number_format = "mm/dd/yyyy"
        ws.cell(row=row, column=7).number_format = "$#,##0.00"

        for column_index in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=column_index).border = Border(bottom=row_separator)


def apply_total_border(ws, row_number: int, last_column: int) -> None:
    thin_gray = Side(style="thin", color="808080")

    for column_index in range(1, last_column + 1):
        ws.cell(row=row_number, column=column_index).border = Border(top=thin_gray)


def write_table_totals(ws, total_row: int, data_start_row: int, data_end_row: int) -> int:
    if data_end_row >= data_start_row:
        quantity_formula = f"=SUM(B{data_start_row}:B{data_end_row})"
        total_formula = f"=SUM(G{data_start_row}:G{data_end_row})"
    else:
        quantity_formula = "=0"
        total_formula = "=0"

    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=quantity_formula).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=total_formula).font = Font(bold=True)

    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right")

    ws.cell(row=total_row, column=2).number_format = "#,##0"
    ws.cell(row=total_row, column=7).number_format = "$#,##0.00"
    apply_total_border(ws, total_row, len(COLUMNS))

    return total_row


def write_ranch_section(ws, start_row: int, ranch_name: str, inventory_df: pd.DataFrame) -> tuple[int, int]:
    header_row = write_table_header(ws, start_row, ranch_name)
    data_start_row = header_row + 1
    next_row = write_inventory_rows(ws, data_start_row, inventory_df)
    data_end_row = next_row - 1

    format_data_rows(ws, data_start_row, data_end_row)

    total_row = next_row if not inventory_df.empty else data_start_row
    write_table_totals(ws, total_row, data_start_row, data_end_row)

    return total_row + 2, total_row


def write_section_title(ws, row_number: int, section_title: str) -> int:
    ws.cell(row=row_number, column=1, value=section_title).font = Font(bold=True, size=13)
    return row_number + 2


def write_location_sections(
    ws, start_row: int, section_title: str, location_inventories: dict[str, pd.DataFrame]
) -> tuple[int, list[int]]:
    current_row = write_section_title(ws, start_row, section_title)
    total_rows = []

    for location, inventory_df in location_inventories.items():
        current_row, total_row = write_ranch_section(ws, current_row, location, inventory_df)
        total_rows.append(total_row)

    return current_row, total_rows


def write_global_total(ws, row_number: int, total_rows: list[int]) -> None:
    quantity_formula = "=" + "+".join(f"B{row}" for row in total_rows) if total_rows else "=0"
    total_formula = "=" + "+".join(f"G{row}" for row in total_rows) if total_rows else "=0"

    ws.cell(row=row_number, column=1, value="TOTAL").font = Font(bold=True, size=14)
    ws.cell(row=row_number, column=2, value=quantity_formula).font = Font(bold=True, size=14)
    ws.cell(row=row_number, column=7, value=total_formula).font = Font(bold=True, size=14)

    ws.cell(row=row_number, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row_number, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row_number, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=row_number, column=2).number_format = "#,##0"
    ws.cell(row=row_number, column=7).number_format = "$#,##0.00"
    apply_total_border(ws, row_number, len(COLUMNS))


def apply_vertical_centering(ws, last_row: int, last_column: int) -> None:
    for row_number in range(1, last_row + 1):
        for column_index in range(1, last_column + 1):
            cell = ws.cell(row=row_number, column=column_index)
            alignment = copy(cell.alignment)
            alignment.vertical = "center"
            cell.alignment = alignment


def apply_print_layout(ws, last_row: int) -> None:
    for row_number in range(1, last_row + 1):
        ws.row_dimensions[row_number].height = 18

    apply_vertical_centering(ws, last_row, len(COLUMNS))
    ws.print_area = f"A1:G{last_row}"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def generate_inventory_report(
    inventories: dict[str, InventoryAssignment],
    output_path: Path | None = None,
    record_sheets: list[InventoryRecordSheet] | None = None,
) -> Path:
    if output_path is None:
        output_path = build_output_path()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Inventory Report"

    apply_sheet_styles(worksheet)
    write_headers(worksheet)

    current_row = 6
    total_rows = []

    for ranch_name, inventory_df in inventories.items():
        if isinstance(inventory_df, dict):
            section_title = RANCH_SECTION_TITLES.get(ranch_name, ranch_name)
            current_row, section_total_rows = write_location_sections(
                worksheet, current_row, section_title, inventory_df
            )
            total_rows.extend(section_total_rows)
        else:
            section_title = RANCH_SECTION_TITLES.get(ranch_name)
            if section_title:
                current_row = write_section_title(worksheet, current_row, section_title)
            current_row, total_row = write_ranch_section(worksheet, current_row, ranch_name, inventory_df)
            total_rows.append(total_row)

    write_global_total(worksheet, current_row, total_rows)
    apply_print_layout(worksheet, current_row)

    if record_sheets is None:
        record_sheets = globals().get("inventory_record_sheets", [])

    for record_sheet in record_sheets:
        write_inventory_record_sheet(workbook, record_sheet)

    workbook.save(output_path)

    return output_path


ranch_dataframes = load_ranch_dataframes()
inventory_assignments = build_inventory_assignments(ranch_dataframes)

# Variables finales para usar en otros scripts.
california_location_inventories = inventory_assignments[CALIFORNIA_RANCH_NAME]
gold_star_inv = california_location_inventories.get("Gold Star Cattle", pd.DataFrame())
vazquez_calf_ranch_inv = california_location_inventories.get("Vazquez Calf Ranch", pd.DataFrame())
la_esperanza_inv = inventory_assignments["La Esperanza Ranch"]
cesar_frias_ranch_inv = inventory_assignments["Cesar Frias Ranch"]
frias_ranch_inv = cesar_frias_ranch_inv
fullmer_cattle_inv = inventory_assignments["Fullmer Cattle"]
inventory_record_sheets = build_inventory_record_sheets(
    ranch_dataframes=ranch_dataframes,
    ranch_section_titles=RANCH_SECTION_TITLES,
    california_ranch_name=CALIFORNIA_RANCH_NAME,
    no_location_label=NO_LOCATION_LABEL,
    california_locations=list(california_location_inventories.keys()),
)


if __name__ == "__main__":
    report_path = generate_inventory_report(inventory_assignments)
    print(f"Reporte creado en: {report_path}")
